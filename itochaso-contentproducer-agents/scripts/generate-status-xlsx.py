#!/usr/bin/env python3
"""
ページ管理ステータス Excel 生成スクリプト
実行: python scripts/generate-status-xlsx.py
出力: pages/{{OUTPUT_FILENAME}}.xlsx

slack-status.json + pages/ スキャン → Excel（アクティブ / アーカイブ / サマリー）
デプロイ済み・否認ページは自動でアーカイブシートへ移動。

各クライアント向けにカスタマイズしてください:
- WP_PAGE_ID と LABEL_MAP を config.yaml で管理する
- SECTIONS をクライアント固有のディレクトリ構成に合わせる
"""

import json
import os
import yaml
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
STATUS_FILE = BASE_DIR / "pages" / "preview" / "slack-status.json"
OUTPUT_FILE = BASE_DIR / "pages" / "preview" / "ページ管理.xlsx"

# config.yaml から読み込む（フォールバック: デフォルト値）
try:
    with open(BASE_DIR / 'config.yaml', 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f) or {}
        WP_PREVIEW_BASE = config.get('wordpress', {}).get('url', 'https://example.com')
        PREVIEW_TOKEN = config.get('wordpress', {}).get('preview_token', 'preview-2026')
except:
    WP_PREVIEW_BASE = "https://example.com"
    PREVIEW_TOKEN = "preview-2026"

# ========================================
# WP_PAGE_ID と LABEL_MAP をカスタマイズしてください
# config.yaml で管理することを推奨します
# ========================================
WP_PAGE_ID = {}
LABEL_MAP = {}

# config.yaml から読み込む場合:
# if 'page_mappings' in config:
#     WP_PAGE_ID = config['page_mappings'].get('wp_page_id', {})
#     LABEL_MAP = config['page_mappings'].get('label_map', {})


# ========================================
# ディレクトリ構成をカスタマイズしてください
# config.yaml で管理することを推奨します
# ========================================
SECTIONS = [
    # {"dir": "pages/preview", "label": "本院", "root_files": []},
    # {"dir": "pages/service", "label": "サービス", "root_files": []},
    # {"dir": "pages/articles", "label": "記事・コラム", "root_files": ["pages/articles-index.html"]},
]

# config.yaml から読み込む場合:
# if 'sections' in config:
#     SECTIONS = config['sections']

EXCLUDE = {"_legacy", "_link-patches", "プレビュー一覧"}

# --- Colors ---
PINK = "B50070"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
HEADER_FILL = PatternFill("solid", fgColor=PINK)
HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=10)
BODY_FONT = Font(name="Arial", size=9)
BODY_FONT_BOLD = Font(name="Arial", size=9, bold=True)
LINK_FONT = Font(name="Arial", size=8, color="0066CC", underline="single")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)

STATUS_COLORS = {
    "未依頼": PatternFill("solid", fgColor="FFF3E0"),
    "確認依頼済": PatternFill("solid", fgColor="E8F5E9"),
    "条件付き承認": PatternFill("solid", fgColor="FFF8E1"),
    "公開承認済": PatternFill("solid", fgColor="E0F2F1"),
    "公開済み": PatternFill("solid", fgColor="E3F2FD"),
    "否認": PatternFill("solid", fgColor="FFEBEE"),
}
STATUS_FONTS = {
    "未依頼": Font(name="Arial", size=9, color="E65100"),
    "確認依頼済": Font(name="Arial", size=9, color="2E7D32"),
    "条件付き承認": Font(name="Arial", size=9, color="F57F17"),
    "公開承認済": Font(name="Arial", size=9, color="00695C", bold=True),
    "公開済み": Font(name="Arial", size=9, color="1565C0", bold=True),
    "否認": Font(name="Arial", size=9, color="C62828"),
}


def load_status():
    try:
        with open(STATUS_FILE) as f:
            return json.load(f)
    except Exception:
        return {"requests": {}, "deploy_plans": {}}


def collect_html(dir_rel):
    full = BASE_DIR / dir_rel
    if not full.exists():
        return []
    results = []
    for p in sorted(full.rglob("*.html")):
        rel = str(p.relative_to(BASE_DIR)).replace("\\", "/")
        if any(ex in rel for ex in EXCLUDE):
            continue
        results.append(rel)
    return results


def get_section(file_path):
    for s in SECTIONS:
        if file_path.startswith(s["dir"]):
            return s["label"]
    return "その他"


def resolve_status(file_path, status_data):
    req = status_data.get("requests", {}).get(file_path)
    if not req:
        return "未依頼", None, None, None

    approval = req.get("approval")
    deployed = req.get("deployed")
    rejected = req.get("rejected")

    if deployed:
        return "公開済み", req.get("deployed_at", ""), req.get("approval_message", ""), req.get("requested_at", "")
    if rejected:
        return "否認", req.get("rejected_at", ""), req.get("rejection_reason", ""), req.get("requested_at", "")
    if approval == "approved":
        return "公開承認済", req.get("approval_at", ""), req.get("approval_message", ""), req.get("requested_at", "")
    if approval == "conditional":
        return "条件付き承認", req.get("approval_at", ""), req.get("approval_message", ""), req.get("requested_at", "")
    return "確認依頼済", req.get("requested_at", ""), None, req.get("requested_at", "")


def wp_preview_url(file_path):
    wp_id = WP_PAGE_ID.get(file_path)
    if wp_id:
        return f"{WP_PREVIEW_BASE}/?page_id={wp_id}&preview_token={PREVIEW_TOKEN}"
    return ""


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_w=8, max_w=45):
    for col_cells in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = min(length + 2, max_w)
        ws.column_dimensions[col_letter].width = max_len


def build_xlsx():
    status_data = load_status()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    all_files = []
    for sec in SECTIONS:
        root = [f for f in sec["root_files"] if (BASE_DIR / f).exists()]
        dir_files = collect_html(sec["dir"])
        all_files.extend(root + dir_files)

    # Classify pages
    active_rows = []
    archive_rows = []

    for fp in all_files:
        label = LABEL_MAP.get(fp, Path(fp).stem)
        section = get_section(fp)
        status, date_str, message, requested_at = resolve_status(fp, status_data)
        preview_url = wp_preview_url(fp)
        date_short = (date_str or "")[:10] if date_str else ""

        row_data = {
            "file": fp,
            "label": label,
            "section": section,
            "status": status,
            "date": date_short,
            "message": (message or "")[:100],
            "preview_url": preview_url,
            "requested_at": (requested_at or "")[:10] if requested_at else "",
        }

        if status in ("公開済み", "否認"):
            archive_rows.append(row_data)
        else:
            active_rows.append(row_data)

    # --- Workbook ---
    wb = Workbook()

    # ========== Sheet 1: アクティブ ==========
    ws1 = wb.active
    ws1.title = "アクティブ"
    ws1.sheet_properties.tabColor = PINK

    headers1 = ["No.", "セクション", "ページ名", "ステータス", "依頼日", "備考", "プレビューURL"]
    ws1.append(headers1)
    style_header(ws1, 1, len(headers1))
    ws1.row_dimensions[1].height = 28
    ws1.freeze_panes = "A2"

    for i, r in enumerate(active_rows, 1):
        row_num = i + 1
        ws1.cell(row=row_num, column=1, value=i).font = BODY_FONT
        ws1.cell(row=row_num, column=2, value=r["section"]).font = BODY_FONT
        ws1.cell(row=row_num, column=3, value=r["label"]).font = BODY_FONT_BOLD
        status_cell = ws1.cell(row=row_num, column=4, value=r["status"])
        status_cell.font = STATUS_FONTS.get(r["status"], BODY_FONT)
        status_cell.fill = STATUS_COLORS.get(r["status"], PatternFill())
        status_cell.alignment = Alignment(horizontal="center")
        ws1.cell(row=row_num, column=5, value=r["date"]).font = BODY_FONT
        ws1.cell(row=row_num, column=6, value=r["message"]).font = BODY_FONT
        if r["preview_url"]:
            url_cell = ws1.cell(row=row_num, column=7, value=r["preview_url"])
            url_cell.font = LINK_FONT
            url_cell.hyperlink = r["preview_url"]
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).border = THIN_BORDER
        if i % 2 == 0:
            for col in range(1, len(headers1) + 1):
                c = ws1.cell(row=row_num, column=col)
                if col != 4:
                    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    ws1.auto_filter.ref = f"A1:G{len(active_rows) + 1}"

    # ========== Sheet 2: アーカイブ ==========
    ws2 = wb.create_sheet("アーカイブ")
    ws2.sheet_properties.tabColor = "1565C0"

    headers2 = ["No.", "セクション", "ページ名", "最終ステータス", "完了日", "依頼日", "備考"]
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    if archive_rows:
        for i, r in enumerate(archive_rows, 1):
            row_num = i + 1
            ws2.cell(row=row_num, column=1, value=i).font = BODY_FONT
            ws2.cell(row=row_num, column=2, value=r["section"]).font = BODY_FONT
            ws2.cell(row=row_num, column=3, value=r["label"]).font = BODY_FONT_BOLD
            status_cell = ws2.cell(row=row_num, column=4, value=r["status"])
            status_cell.font = STATUS_FONTS.get(r["status"], BODY_FONT)
            status_cell.fill = STATUS_COLORS.get(r["status"], PatternFill())
            status_cell.alignment = Alignment(horizontal="center")
            ws2.cell(row=row_num, column=5, value=r["date"]).font = BODY_FONT
            ws2.cell(row=row_num, column=6, value=r["requested_at"]).font = BODY_FONT
            ws2.cell(row=row_num, column=7, value=r["message"]).font = BODY_FONT
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row_num, column=col).border = THIN_BORDER
        ws2.auto_filter.ref = f"A1:G{len(archive_rows) + 1}"
    else:
        ws2.cell(row=2, column=1, value="（まだアーカイブはありません）").font = Font(name="Arial", size=9, color="999999")

    # ========== Sheet 3: サマリー ==========
    ws3 = wb.create_sheet("サマリー")
    ws3.sheet_properties.tabColor = "00695C"

    ws3.cell(row=1, column=1, value="操レディスホスピタル ページ管理サマリー").font = Font(name="Arial", size=14, bold=True, color=PINK)
    ws3.merge_cells("A1:D1")
    ws3.cell(row=2, column=1, value=f"更新: {now}").font = Font(name="Arial", size=9, color="888888")

    # Status counts using formulas
    statuses = ["未依頼", "確認依頼済", "条件付き承認", "公開承認済", "公開済み", "否認"]
    ws3.cell(row=4, column=1, value="ステータス").font = HEADER_FONT
    ws3.cell(row=4, column=1).fill = HEADER_FILL
    ws3.cell(row=4, column=2, value="件数").font = HEADER_FONT
    ws3.cell(row=4, column=2).fill = HEADER_FILL
    ws3.row_dimensions[4].height = 24

    active_range = f"アクティブ!D2:D{max(len(active_rows) + 1, 2)}"
    archive_range = f"アーカイブ!D2:D{max(len(archive_rows) + 1, 2)}"

    for i, s in enumerate(statuses):
        row = 5 + i
        cell_label = ws3.cell(row=row, column=1, value=s)
        cell_label.font = STATUS_FONTS.get(s, BODY_FONT)
        cell_label.fill = STATUS_COLORS.get(s, PatternFill())
        if s in ("公開済み", "否認"):
            ws3.cell(row=row, column=2).value = f'=COUNTIF({archive_range},"{s}")'
        else:
            ws3.cell(row=row, column=2).value = f'=COUNTIF({active_range},"{s}")'
        ws3.cell(row=row, column=2).font = Font(name="Arial", size=11, bold=True)
        ws3.cell(row=row, column=2).alignment = Alignment(horizontal="center")

    total_row = 5 + len(statuses)
    ws3.cell(row=total_row, column=1, value="合計").font = Font(name="Arial", size=10, bold=True)
    ws3.cell(row=total_row, column=2).value = f"=SUM(B5:B{total_row - 1})"
    ws3.cell(row=total_row, column=2).font = Font(name="Arial", size=11, bold=True)
    ws3.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    for col in range(1, 3):
        ws3.cell(row=total_row, column=col).border = Border(top=Side(style="medium", color="333333"))

    # Section breakdown
    sec_start = total_row + 2
    ws3.cell(row=sec_start, column=1, value="セクション別").font = HEADER_FONT
    ws3.cell(row=sec_start, column=1).fill = HEADER_FILL
    ws3.cell(row=sec_start, column=2, value="アクティブ").font = HEADER_FONT
    ws3.cell(row=sec_start, column=2).fill = HEADER_FILL
    ws3.cell(row=sec_start, column=3, value="アーカイブ").font = HEADER_FONT
    ws3.cell(row=sec_start, column=3).fill = HEADER_FILL
    ws3.row_dimensions[sec_start].height = 24

    section_names = [s["label"] for s in SECTIONS]
    active_sec_range = f"アクティブ!B2:B{max(len(active_rows) + 1, 2)}"
    archive_sec_range = f"アーカイブ!B2:B{max(len(archive_rows) + 1, 2)}"
    for i, sn in enumerate(section_names):
        row = sec_start + 1 + i
        ws3.cell(row=row, column=1, value=sn).font = BODY_FONT_BOLD
        ws3.cell(row=row, column=2).value = f'=COUNTIF({active_sec_range},"{sn}")'
        ws3.cell(row=row, column=2).font = BODY_FONT
        ws3.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws3.cell(row=row, column=3).value = f'=COUNTIF({archive_sec_range},"{sn}")'
        ws3.cell(row=row, column=3).font = BODY_FONT
        ws3.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    # Auto-width all sheets
    auto_width(ws1)
    auto_width(ws2)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14

    # Override specific column widths for sheet 1
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["G"].width = 50

    wb.save(str(OUTPUT_FILE))
    print(f"✅ Excel生成完了: {OUTPUT_FILE}")
    print(f"   アクティブ: {len(active_rows)}件")
    print(f"   アーカイブ: {len(archive_rows)}件")
    print(f"   更新日時: {now}")


if __name__ == "__main__":
    build_xlsx()
